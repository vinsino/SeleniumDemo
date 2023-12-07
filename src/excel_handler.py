import os
import openpyxl
from openpyxl.utils import column_index_from_string

# Read data from an existing Excel file
# def read_excel(file_path):
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active

#     # Assuming data is in the first column
#     column_data = [cell.value for cell in sheet['A']]

#     workbook.close()
#     return column_data

# Read data from an existing Excel file
def read_excel(file_path):
    
    # columns_to_read = ['B', 'C', 'F', 'G']  # Replace with your actual column letters or indices
    columns_to_read = ['B', 'F', 'G']  # Replace with your actual column letters or indices
    index_to_read = list(map(lambda x: column_index_from_string(x) - 1, columns_to_read))

    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Initialize an empty list to store the data
    data = []

    # Iterate over rows and columns
    for row in sheet.iter_rows(min_row=3, values_only=True):  # Assuming data starts from the second row
        row_data = [row[idx] for idx in index_to_read]
        if None in row_data:
            break
        data.append(row_data)

    # Display the selected columns
    # print("Data from selected columns:")
    # for row in data:
    #     print(row)

    workbook.close()
    return data


# Write data to a new Excel file
def write_excel(file_path, data):
    
    # workbook = openpyxl.Workbook()
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    columns_to_write = ['C', 'D', 'E', 'H', 'I', 'J', 'K', 'L', 'M', 'N']  # Replace with your actual column letters or indices
    index_to_write = list(map(column_index_from_string, columns_to_write))

    # # Write data to the first column
    # for index, value in enumerate(data, start=1):
    #     sheet.cell(row=index, column=1, value=value)    

    # Write the data to the worksheet
    start_row = 3

    for row_index, row_data in enumerate(data, start=0):
        for col_index, value in enumerate(row_data, start=0):
            sheet.cell(row=row_index + start_row, column=index_to_write[col_index], value=value)


    # Split the original filepath into the directory and the filename
    directory, filename = os.path.split(file_path)
    output_directory = './output'
    # Construct the new file path
    new_filepath = os.path.join(output_directory, f'{os.path.splitext(filename)[0]}_done{os.path.splitext(filename)[1]}')


    workbook.save(new_filepath)
    workbook.close()

# # Example usage
# input_file_path = 'input.xlsx'
# output_file_path = 'output.xlsx'

# # Read data from existing Excel file
# data_to_write = read_excel(input_file_path)

# # Modify or process the data as needed

# # Write data to a new Excel file
# write_excel(output_file_path, data_to_write)
